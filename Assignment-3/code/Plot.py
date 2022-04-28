import matplotlib.pyplot as plt
import openpyxl

# Accesing tables
wb1=openpyxl.load_workbook("C:\\Users\\arisa\\Documents\\Book1.xlsx")
wb2=openpyxl.load_workbook("C:\\Users\\arisa\\Documents\\Book2.xlsx")
wb3=openpyxl.load_workbook("C:\\Users\\arisa\\Documents\\Book3.xlsx")
sh1=wb1['Sheet1']
sh2=wb2['Sheet1']
sh3=wb3['Sheet1']

# Data from question
"""  __________________________________________
        Age(in years)   |  Number of childrens 
            1-2         |         5
            2-3         |         3
            3-5         |         6
            5-7         |         12
            7-10        |         9
           10-15        |         10
           15-17        |         4
     ___________________|_______________________
"""
row=sh1.max_row

# Reading data from table 1 given in quetion from Table_1.xlsx
l_val=[]
for i in range(2,row+1):
    l_val.append(int(sh1.cell(i,1).value))
r_val=[]
for i in range(2,row+1):
    r_val.append(int(sh1.cell(i,2).value))
freq=[]
for i in range(2,row+1):
    freq.append(int(sh1.cell(i,4).value))
    
"""
Adjusted Frequency =(Frequency×Required size of interval)/Current size of interval
Required size of interval = minimum size of interval
"""

# Calculating Adjusted frequency
adjusted_freq = []
minimum_interval = min(r_val)-min(l_val)
for i in range(len(freq)):
    adjusted_freq.append((freq[i]*minimum_interval)//(r_val[i]-l_val[i]))
    
# Writing data in Table_2.xlsx    
for i in range(2,row+1):
    sh2.cell(row=i,column=2,value=freq[i-2])
for i in range(2,row+1):
    sh2.cell(row=i,column=4,value=adjusted_freq[i-2])
wb2.save("C:\\Users\\arisa\\Documents\\Book2.xlsx")

# Creating required list 'a' for plotting histogram  
a=[]
for i in range(len(adjusted_freq)):
    for j in range(adjusted_freq[i]):
        a.append((l_val[i]+r_val[i])/2)
        
# Writing data in Table_3.xlsx
for i in range(2,row+1):
    sh3.cell(row=i,column=2,value=adjusted_freq[i-2])
wb3.save("C:\\Users\\arisa\\Documents\\Book3.xlsx")

# Lables
plt.xlabel('No. of childrens')
plt.ylabel('Age(in years)')

# Plotting graph
plt.title('Analysis of childern of various age group playing in playground')
plt.hist(a,bins = [1,2,3,5,7,10,15,17],facecolor='r',alpha=0.9,rwidth=0.97)
plt.show()
